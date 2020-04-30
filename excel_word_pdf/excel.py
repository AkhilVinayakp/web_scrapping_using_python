import openpyxl as xl
# opening the workbook via load_workbook() method
wb = xl.load_workbook('example.xlsx')
print(type(wb)) #type of the workbook will print
#knowing the names of the sheet
sheet_names =next(iter(wb.get_sheet_names()),None) # only gets one sheet name since next is being used
print(sheet_names)
sheet = wb.get_sheet_by_name(sheet_names) # geting the sheet object
print(sheet['A1'].value)
#geting the the cell object using cell() method
cell = sheet.cell(row=1,column=1)
print(cell.value)
